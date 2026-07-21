"""
Export module — Excel (.xlsx)

Public interface:
    generate_excel(results: list) -> bytes

Exports ALL results: shortlisted (rank 1–N) first, then unranked below a divider.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def generate_excel(results: list) -> bytes:
    """
    Generate an Excel workbook from a list of shortlisted search results.

    Args:
        results: List of SearchResult ORM objects (all results for the search).

    Returns:
        Raw bytes of the .xlsx file ready to stream as a download.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["Rank", "Name / Title", "URL", "Relevance Score", "Reasoning"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Shortlisted results ───────────────────────────────────────────────────
    shortlist = sorted([r for r in results if r.final_rank is not None], key=lambda r: r.final_rank)
    other = [r for r in results if r.final_rank is None]

    for result in shortlist:
        ws.append([result.final_rank, result.title or "", result.url, result.relevance_score, result.relevance_reasoning or ""])

    # ── Divider + unranked results ────────────────────────────────────────────
    if other:
        divider_row = ws.max_row + 1
        divider_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        cell = ws.cell(row=divider_row, column=1, value="Other Results (not in shortlist)")
        cell.font = Font(bold=True, color="375623")
        cell.fill = divider_fill
        for col in range(2, 6):
            ws.cell(row=divider_row, column=col).fill = divider_fill
        for result in other:
            ws.append(["", result.title or "", result.url, result.relevance_score, result.relevance_reasoning or ""])

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [8, 45, 60, 18, 80]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    ws.row_dimensions[1].height = 20

    # ── Serialise to bytes ────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
